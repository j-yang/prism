"""Excel Writer for Meta Output.

Generates formatted Excel files with multiple sheets.
"""

import json
from pathlib import Path
from typing import Any, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from prism.meta.definitions.models import MetaDefinitions


def get_value(obj: Any, key: str, default: Any = None) -> Any:
    """Get value from dict or Pydantic model."""
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


HEADER_FILLS = {
    "study_config": PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    ),
    "params": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    "silver_variables": PatternFill(
        start_color="FFC000", end_color="FFC000", fill_type="solid"
    ),
    "platinum": PatternFill(
        start_color="7030A0", end_color="7030A0", fill_type="solid"
    ),
    "gold_statistics": PatternFill(
        start_color="ED7D31", end_color="ED7D31", fill_type="solid"
    ),
}

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

NORMAL_FONT = Font(size=10)

BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

NEEDS_REVIEW_FILL = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)


class MetaExcelWriter:
    """Write generated metadata to formatted Excel file."""

    def __init__(self, output_path: str):
        self.output_path = Path(output_path)
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def write_specs(self, specs: List[MetaDefinitions], study_info: dict = None):
        """Write all meta definitions to Excel sheets.

        Args:
            specs: List of MetaDefinitions objects
            study_info: Optional study info dict
        """
        all_platinum = []
        all_silver = []
        all_gold = []
        all_params = []
        study_config = {}
        confidence_notes = []

        for spec in specs:
            all_platinum.extend(spec.platinum_deliverables)
            all_silver.extend(spec.silver_variables)
            all_gold.extend(spec.gold_statistics)
            all_params.extend(spec.params)
            confidence_notes.extend(spec.confidence_notes)

        self._write_study_config(study_config, study_info or {})
        self._write_params(all_params)
        self._write_silver_variables(all_silver)
        self._write_platinum(all_platinum)
        self._write_gold_statistics(all_gold)

        if confidence_notes:
            self._write_confidence_notes(confidence_notes)

    def _write_study_config(self, study_config: dict, study_info: dict):
        """Write study_config sheet."""
        ws = self.wb.create_sheet("study_config")

        rows = [["category", "key", "value", "description"]]

        if study_info:
            for key, value in study_info.items():
                rows.append(["study_info", key, str(value), ""])

        for pop in study_config.get("populations", []):
            rows.append(
                [
                    "population",
                    pop.get("name", ""),
                    pop.get("selection", ""),
                    pop.get("description", ""),
                ]
            )

        for period in study_config.get("event_periods", []):
            rows.append(
                [
                    "event_period",
                    period.get("name", ""),
                    period.get("selection", ""),
                    period.get("description", ""),
                ]
            )

        self._write_table(ws, rows, "study_config")

    def _write_params(self, params: list):
        """Write params sheet."""
        ws = self.wb.create_sheet("params")

        if not params:
            self._write_table(
                ws,
                [
                    [
                        "paramcd",
                        "parameter",
                        "category",
                        "unit",
                        "source_form",
                        "used_in",
                    ]
                ],
                "params",
            )
            return

        rows = []
        seen = set()

        for p in params:
            paramcd = get_value(p, "paramcd", "")
            if paramcd and paramcd not in seen:
                seen.add(paramcd)
                used_in = get_value(p, "used_in", "")
                if isinstance(used_in, list):
                    used_in = ", ".join(str(x) for x in used_in)
                rows.append(
                    [
                        paramcd,
                        get_value(p, "parameter", ""),
                        get_value(p, "category", ""),
                        get_value(p, "unit", ""),
                        get_value(p, "source_form", ""),
                        used_in,
                    ]
                )

        rows.insert(
            0, ["paramcd", "parameter", "category", "unit", "source_form", "used_in"]
        )
        self._write_table(ws, rows, "params")

    def _write_silver_variables(self, variables: list):
        """Write silver_variables sheet with meta-compatible columns."""
        ws = self.wb.create_sheet("silver_variables")

        rows = [
            [
                "var_name",
                "var_label",
                "schema",
                "data_type",
                "description",
                "source_vars",
                "transformation",
                "transformation_type",
                "param_ref",
                "confidence",
                "used_in",
            ]
        ]

        seen = set()
        for v in variables:
            var_name = get_value(v, "var_name", "")
            if var_name and var_name not in seen:
                seen.add(var_name)
                source_vars = get_value(v, "source_vars", "")
                if isinstance(source_vars, list):
                    source_vars = ", ".join(str(x) for x in source_vars)
                used_in = get_value(v, "used_in", "")
                if isinstance(used_in, list):
                    used_in = ", ".join(str(x) for x in used_in)
                var_label = get_value(v, "var_label", "") or get_value(v, "label", "")
                transformation = get_value(v, "transformation", "") or get_value(
                    v, "derivation", ""
                )
                rows.append(
                    [
                        var_name,
                        var_label,
                        get_value(v, "schema", ""),
                        get_value(v, "data_type", ""),
                        get_value(v, "description", ""),
                        source_vars,
                        transformation,
                        get_value(v, "transformation_type", "direct"),
                        get_value(v, "param_ref", ""),
                        get_value(v, "confidence", "medium"),
                        used_in,
                    ]
                )

        self._write_table(ws, rows, "silver_variables")

        self._apply_conditional_format(ws, len(rows), "confidence", ["low"])

    def _write_platinum(self, deliverables: list):
        """Write platinum sheet."""
        ws = self.wb.create_sheet("platinum")

        rows = [["deliverable_id", "deliverable_type", "title", "population", "schema"]]

        seen = set()
        for d in deliverables:
            deliv_id = get_value(d, "deliverable_id", "")
            if deliv_id and deliv_id not in seen:
                seen.add(deliv_id)
                rows.append(
                    [
                        deliv_id,
                        get_value(d, "deliverable_type", ""),
                        get_value(d, "title", ""),
                        get_value(d, "population", ""),
                        get_value(d, "schema", ""),
                    ]
                )

        self._write_table(ws, rows, "platinum")

    def _write_gold_statistics(self, stats: list):
        """Write gold_statistics sheet."""
        ws = self.wb.create_sheet("gold_statistics")

        rows = [
            [
                "deliverable_id",
                "row_label",
                "element_type",
                "element_id",
                "selection",
                "statistics",
                "group_by",
            ]
        ]

        for s in stats:
            group_by = get_value(s, "group_by", "")
            if isinstance(group_by, list):
                group_by = ", ".join(str(x) for x in group_by)
            statistics = get_value(s, "statistics", "")
            if isinstance(statistics, dict):
                statistics = json.dumps(statistics)
            rows.append(
                [
                    get_value(s, "deliverable_id", ""),
                    get_value(s, "row_label", ""),
                    get_value(s, "element_type", ""),
                    get_value(s, "element_id", ""),
                    get_value(s, "selection", ""),
                    statistics or "",
                    group_by,
                ]
            )

        self._write_table(ws, rows, "gold_statistics")

    def _write_confidence_notes(self, notes: list):
        """Write confidence_notes sheet for items needing review."""
        ws = self.wb.create_sheet("review_needed")

        rows = [["deliverable_id", "note"]]
        rows.extend(
            [[get_value(n, "deliverable", ""), get_value(n, "note", "")] for n in notes]
        )

        self._write_table(ws, rows, "study_config")

        for row in range(2, len(rows) + 1):
            for col in range(1, 3):
                ws.cell(row=row, column=col).fill = NEEDS_REVIEW_FILL

    def _write_table(self, ws, rows: list, sheet_type: str):
        """Write a table with formatting."""
        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if row_idx == 1:
                    cell.fill = HEADER_FILLS.get(
                        sheet_type, HEADER_FILLS["study_config"]
                    )
                    cell.font = HEADER_FONT
                else:
                    cell.font = NORMAL_FONT

                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = BORDER

        self._auto_fit_columns(ws, rows)

        ws.freeze_panes = "A2"

    def _auto_fit_columns(self, ws, rows: list):
        """Auto-fit column widths."""
        if not rows:
            return

        col_widths = {}

        for row in rows:
            for col_idx, value in enumerate(row, 1):
                value_str = str(value) if value else ""
                width = min(max(len(value_str), 10), 60)
                col_widths[col_idx] = max(col_widths.get(col_idx, 0), width)

        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

    def _apply_conditional_format(
        self, ws, row_count: int, col_name: str, highlight_values: list
    ):
        """Apply conditional formatting to highlight specific values."""
        headers = [cell.value for cell in ws[1]]

        try:
            col_idx = headers.index(col_name) + 1
        except ValueError:
            return

        for row in range(2, row_count + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value in highlight_values:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = NEEDS_REVIEW_FILL

    def save(self):
        """Save the workbook."""
        self.wb.save(self.output_path)
        return str(self.output_path)


def write_meta_excel(
    specs: List[MetaDefinitions], output_path: str, study_info: dict = None
) -> str:
    """Convenience function to write meta definitions to Excel.

    Args:
        specs: List of MetaDefinitions objects
        output_path: Output file path
        study_info: Optional study info dict

    Returns:
        Path to saved file
    """
    writer = MetaExcelWriter(output_path)
    writer.write_specs(specs, study_info)
    return writer.save()


write_spec_excel = write_meta_excel
SpecExcelWriter = MetaExcelWriter
