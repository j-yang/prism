"""ALS Variable Matcher.

Matches spec variables to ALS fields using LLM for semantic matching.
"""

import json
from dataclasses import dataclass
from typing import Optional, List, Dict

from prism.agent import call_deepseek
from prism.spec.templates import format_prompt, TEMPLATE_MATCH_ALS


@dataclass
class MatchResult:
    """Result of ALS variable matching."""

    var_name: str
    matched: bool
    als_field: Optional[str]
    als_label: Optional[str]
    confidence: str  # 'high', 'medium', 'low', 'n/a'
    notes: str


class ALSMatcher:
    """Match variables to ALS fields using LLM."""

    def __init__(self, als_dict: Dict[str, dict]):
        """
        Args:
            als_dict: Dictionary mapping field names to their info
                     {field_name: {label, data_type, form, ...}}
        """
        self.als_dict = als_dict

    def match_variables(self, variables: List[dict]) -> List[MatchResult]:
        """Match a list of variables to ALS fields.

        Args:
            variables: List of variable dicts with 'var_name' and 'label'

        Returns:
            List of MatchResult objects
        """
        als_json = json.dumps(self.als_dict, indent=2, ensure_ascii=False)[:8000]
        vars_json = json.dumps(variables, indent=2, ensure_ascii=False)[:4000]

        prompt = format_prompt(
            TEMPLATE_MATCH_ALS, als_dict_json=als_json, variables_json=vars_json
        )

        response = call_deepseek(prompt, temperature=0.1)

        if response:
            try:
                results = json.loads(response)
                return [MatchResult(**r) for r in results]
            except json.JSONDecodeError:
                pass

        return [
            MatchResult(
                var_name=v.get("var_name", ""),
                matched=False,
                als_field=None,
                als_label=None,
                confidence="low",
                notes="LLM matching failed, manual review needed",
            )
            for v in variables
        ]

    def find_by_label_similarity(self, label: str) -> Optional[dict]:
        """Find ALS field by label similarity (simple keyword match).

        Used as fallback when LLM is unavailable.
        """
        label_lower = label.lower()
        keywords = set(label_lower.replace("(", " ").replace(")", " ").split())

        best_match = None
        best_score = 0

        for field_name, field_info in self.als_dict.items():
            field_label = field_info.get("label", "").lower()
            field_keywords = set(field_label.split())

            overlap = len(keywords & field_keywords)
            if overlap > best_score:
                best_score = overlap
                best_match = {
                    "field": field_name,
                    "label": field_info.get("label"),
                    "score": overlap,
                }

        return best_match if best_score > 0 else None


def load_als_dict(als_path: str) -> Dict[str, dict]:
    """Load ALS dictionary from Excel file.

    Args:
        als_path: Path to ALS Excel file

    Returns:
        Dictionary mapping field names to their info
    """
    import openpyxl

    wb = openpyxl.load_workbook(als_path, read_only=True, data_only=True)

    als_dict = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        header = None
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                header = [
                    str(h).lower().replace(" ", "_") if h else f"col_{i}"
                    for i, h in enumerate(row)
                ]
                continue

            if not any(row):
                continue

            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(header):
                    row_dict[header[i]] = str(cell) if cell else ""

            field_name = (
                row_dict.get("field_oid")
                or row_dict.get("field_name")
                or row_dict.get("variable")
            )
            if field_name:
                form_name = (
                    row_dict.get("form_oid") or row_dict.get("form") or sheet_name
                )
                full_name = f"{form_name}.{field_name}" if form_name else field_name

                als_dict[full_name] = {
                    "field": field_name,
                    "form": form_name,
                    "label": row_dict.get("label") or row_dict.get("field_label") or "",
                    "data_type": row_dict.get("data_type")
                    or row_dict.get("type")
                    or "",
                    "codelist": row_dict.get("codelist")
                    or row_dict.get("code_list")
                    or "",
                }

    wb.close()
    return als_dict


def match_als_variables(variables: List[dict], als_path: str) -> List[MatchResult]:
    """Convenience function to match variables to ALS.

    Args:
        variables: List of variable dicts
        als_path: Path to ALS Excel file

    Returns:
        List of MatchResult objects
    """
    als_dict = load_als_dict(als_path)
    matcher = ALSMatcher(als_dict)
    return matcher.match_variables(variables)
