"""Excel Writer for Spec Output.

Generates formatted Excel files with multiple sheets.
"""

import json
from dataclasses import asdict
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from prism.spec.generator import GeneratedSpec


HEADER_FILLS = {
    "study_config": PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    ),
    "params": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    "silver_variables": PatternFill(
        start_color="FFC000", end_color="FFC000", fill_type="solid"
    ),
    "platinum": PatternFill(
        start_color="7030A0", end_color="7030A0", fill_type="solid"
    ),
    "gold_statistics": PatternFill(
        start_color="ED7D31", end_color="ED7D31", fill_type="solid"
    ),
}

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

NORMAL_FONT = Font(size=10)

BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

NEEDS_REVIEW_FILL = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)


class SpecExcelWriter:
    """Write generated specs to formatted Excel file."""

    def __init__(self, output_path: str):
        self.output_path = Path(output_path)
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def write_specs(self, specs: List[GeneratedSpec], study_info: dict = None):
        """Write all specs to Excel sheets.

        Args:
            specs: List of GeneratedSpec objects
            study_info: Optional study info dict
        """
        all_platinum = []
        all_silver = []
        all_gold = []
        all_params = []
        study_config = {}
        confidence_notes = []

        for spec in specs:
            if spec.platinum:
                all_platinum.append(spec.platinum)
            all_silver.extend(spec.silver_variables)
            all_gold.extend(spec.gold_statistics)
            all_params.extend(spec.params)
            if spec.study_config:
                study_config.update(spec.study_config)
            confidence_notes.extend(
                [
                    {"deliverable": spec.deliverable_id, "note": n}
                    for n in spec.confidence_notes
                ]
            )

        self._write_study_config(study_config, study_info or {})
        self._write_params(all_params)
        self._write_silver_variables(all_silver)
        self._write_platinum(all_platinum)
        self._write_gold_statistics(all_gold)

        if confidence_notes:
            self._write_confidence_notes(confidence_notes)

    def _write_study_config(self, study_config: dict, study_info: dict):
        """Write study_config sheet."""
        ws = self.wb.create_sheet("study_config")

        rows = [["category", "key", "value", "description"]]

        if study_info:
            for key, value in study_info.items():
                rows.append(["study_info", key, str(value), ""])

        for pop in study_config.get("populations", []):
            rows.append(
                [
                    "population",
                    pop.get("name", ""),
                    pop.get("selection", ""),
                    pop.get("description", ""),
                ]
            )

        for period in study_config.get("event_periods", []):
            rows.append(
                [
                    "event_period",
                    period.get("name", ""),
                    period.get("selection", ""),
                    period.get("description", ""),
                ]
            )

        self._write_table(ws, rows, "study_config")

    def _write_params(self, params: list):
        """Write params sheet."""
        ws = self.wb.create_sheet("params")

        if not params:
            self._write_table(
                ws,
                [["paramcd", "parameter", "category", "unit", "source_form"]],
                "params",
            )
            return

        rows = []
        seen = set()

        for p in params:
            paramcd = p.get("paramcd", "")
            if paramcd and paramcd not in seen:
                seen.add(paramcd)
                rows.append(
                    [
                        paramcd,
                        p.get("parameter", ""),
                        p.get("category", ""),
                        p.get("unit", ""),
                        p.get("source_form", ""),
                        p.get("visits", ""),
                    ]
                )

        rows.insert(
            0, ["paramcd", "parameter", "category", "unit", "source_form", "visits"]
        )
        self._write_table(ws, rows, "params")

    def _write_silver_variables(self, variables: list):
        """Write silver_variables sheet."""
        ws = self.wb.create_sheet("silver_variables")

        rows = [
            [
                "var_name",
                "schema",
                "label",
                "data_type",
                "derivation",
                "source_vars",
                "confidence",
                "used_in",
            ]
        ]

        seen = set()
        for v in variables:
            var_name = v.get("var_name", "")
            if var_name and var_name not in seen:
                seen.add(var_name)
                source_vars = v.get("source_vars", "")
                if isinstance(source_vars, list):
                    source_vars = ", ".join(str(x) for x in source_vars)
                rows.append(
                    [
                        var_name,
                        v.get("schema", ""),
                        v.get("label", ""),
                        v.get("data_type", ""),
                        v.get("derivation", ""),
                        source_vars,
                        v.get("confidence", "medium"),
                        v.get("used_in", ""),
                    ]
                )

        self._write_table(ws, rows, "silver_variables")

        self._apply_conditional_format(ws, len(rows), "confidence", ["low"])

    def _write_platinum(self, deliverables: list):
        """Write platinum sheet."""
        ws = self.wb.create_sheet("platinum")

        rows = [["deliverable_id", "deliverable_type", "title", "population", "schema"]]

        seen = set()
        for d in deliverables:
            deliv_id = d.get("deliverable_id", "")
            if deliv_id and deliv_id not in seen:
                seen.add(deliv_id)
                rows.append(
                    [
                        deliv_id,
                        d.get("deliverable_type", ""),
                        d.get("title", ""),
                        d.get("population", ""),
                        d.get("schema", ""),
                    ]
                )

        self._write_table(ws, rows, "platinum")

    def _write_gold_statistics(self, stats: list):
        """Write gold_statistics sheet."""
        ws = self.wb.create_sheet("gold_statistics")

        rows = [
            [
                "deliverable_id",
                "row_label",
                "element_type",
                "element_id",
                "selection",
                "statistics",
                "group_by",
            ]
        ]

        for s in stats:
            group_by = s.get("group_by", "")
            if isinstance(group_by, list):
                group_by = ", ".join(str(x) for x in group_by)
            rows.append(
                [
                    s.get("deliverable_id", ""),
                    s.get("row_label", ""),
                    s.get("element_type", ""),
                    s.get("element_id", ""),
                    s.get("selection", ""),
                    json.dumps(s.get("statistics", {}))
                    if isinstance(s.get("statistics"), dict)
                    else str(s.get("statistics", "")),
                    group_by,
                ]
            )

        self._write_table(ws, rows, "gold_statistics")

    def _write_confidence_notes(self, notes: list):
        """Write confidence_notes sheet for items needing review."""
        ws = self.wb.create_sheet("review_needed")

        rows = [["deliverable_id", "note"]]
        rows.extend([[n.get("deliverable", ""), n.get("note", "")] for n in notes])

        self._write_table(ws, rows, "study_config")

        for row in range(2, len(rows) + 1):
            for col in range(1, 3):
                ws.cell(row=row, column=col).fill = NEEDS_REVIEW_FILL

    def _write_table(self, ws, rows: list, sheet_type: str):
        """Write a table with formatting."""
        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if row_idx == 1:
                    cell.fill = HEADER_FILLS.get(
                        sheet_type, HEADER_FILLS["study_config"]
                    )
                    cell.font = HEADER_FONT
                else:
                    cell.font = NORMAL_FONT

                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = BORDER

        self._auto_fit_columns(ws, rows)

        ws.freeze_panes = "A2"

    def _auto_fit_columns(self, ws, rows: list):
        """Auto-fit column widths."""
        if not rows:
            return

        col_widths = {}

        for row in rows:
            for col_idx, value in enumerate(row, 1):
                value_str = str(value) if value else ""
                width = min(max(len(value_str), 10), 60)
                col_widths[col_idx] = max(col_widths.get(col_idx, 0), width)

        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

    def _apply_conditional_format(
        self, ws, row_count: int, col_name: str, highlight_values: list
    ):
        """Apply conditional formatting to highlight specific values."""
        headers = [cell.value for cell in ws[1]]

        try:
            col_idx = headers.index(col_name) + 1
        except ValueError:
            return

        for row in range(2, row_count + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value in highlight_values:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = NEEDS_REVIEW_FILL

    def save(self):
        """Save the workbook."""
        self.wb.save(self.output_path)
        return str(self.output_path)


def write_spec_excel(
    specs: List[GeneratedSpec], output_path: str, study_info: dict = None
) -> str:
    """Convenience function to write specs to Excel.

    Args:
        specs: List of GeneratedSpec objects
        output_path: Output file path
        study_info: Optional study info dict

    Returns:
        Path to saved file
    """
    writer = SpecExcelWriter(output_path)
    writer.write_specs(specs, study_info)
    return writer.save()
